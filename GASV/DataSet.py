import lasio
import numpy as np
import pandas as pd
import openpyxl
from matplotlib import pyplot as plt
import seaborn as sns
import lasio
import warnings
import random
def dataset_lab(b):
    workbook_1 = openpyxl.load_workbook('./井/WY1.xlsx')
    worksheet_1 = workbook_1.get_sheet_by_name("GASV")
    fileName_1 = './井/weiye1jgw.las'
    las_1 = lasio.read(fileName_1)
    Aim_1 = np.array([item.value for item in list(worksheet_1.columns)[2]]).reshape(-1, 1)
    df_Aim_1 = pd.DataFrame(Aim_1, columns=['GASV'])
    las_1_data_depth = las_1.data[:, 0]
    las_1_data = las_1.data[:, ]
    Para_DTS = np.zeros((Aim_1.shape[0], 1))
    Para_RD = np.zeros((Aim_1.shape[0], 1))
    Para_DT = np.zeros((Aim_1.shape[0], 1))
    Para_GASV_1 = np.zeros((Aim_1.shape[0], 1))
    Para_GR = np.zeros((Aim_1.shape[0], 1))
    Para_RHOB = np.zeros((Aim_1.shape[0], 1))
    Para_CNL = np.zeros((Aim_1.shape[0], 1))
    depth_max = np.array([item.value for item in list(worksheet_1.columns)[1]])
    depth = ((depth_max * 10).astype(int))/10
    for i in range(Aim_1.shape[0]):
        for j in range(las_1_data_depth.shape[0]):
            if (depth[i] == las_1_data_depth[j]):
                Para_DTS[i][0] = las_1_data.data[j, 3]
                Para_RD[i][0] = las_1_data.data[j, 5]
                Para_DT[i][0] = las_1_data.data[j, 7]
                Para_GASV_1[i][0] = las_1_data.data[j, 9]
                Para_GR[i][0] = las_1_data.data[j, 10]
                Para_RHOB[i][0] = las_1_data.data[j, 11]
                Para_CNL[i][0] = las_1_data.data[j, 13]
    data_1 = np.hstack((Para_DTS, Para_DT, Para_RHOB, Para_GR, Aim_1, Para_GASV_1))
    print(data_1.shape)

    workbook_29 = openpyxl.load_workbook('./井/WY29.xlsx')
    worksheet_29 = workbook_29.get_sheet_by_name("GASV")
    fileName_29 = './井/wy29jgw.las'
    las_29 = lasio.read(fileName_29)
    Aim_29 = (np.array([item.value for item in list(worksheet_29.columns)[6]]).reshape(-1, 1)) + np.array([item.value for item in list(worksheet_29.columns)[7]]).reshape(-1, 1)
    df_Aim_29 = pd.DataFrame(Aim_29, columns=['GASV'])
    las_29_data_depth = las_29.data[:, 0]
    las_29_data = las_29.data[:, ]
    Para_DTS = np.zeros((Aim_29.shape[0], 1))
    Para_RD = np.zeros((Aim_29.shape[0], 1))
    Para_DT = np.zeros((Aim_29.shape[0], 1))
    Para_GASV_29 = np.zeros((Aim_29.shape[0], 1))
    Para_GR = np.zeros((Aim_29.shape[0], 1))
    Para_RHOB = np.zeros((Aim_29.shape[0], 1))
    Para_CNL = np.zeros((Aim_29.shape[0], 1))
    depth_max = np.array([item.value for item in list(worksheet_29.columns)[1]])
    depth = ((depth_max * 10).astype(int)) / 10
    for i in range(Aim_29.shape[0]):
        for j in range(las_29_data_depth.shape[0]):
            if (depth[i] == las_29_data_depth[j]):
                Para_DTS[i][0] = las_29_data.data[j, 3]
                Para_RD[i][0] = las_29_data.data[j, 5]
                Para_DT[i][0] = las_29_data.data[j, 7]
                Para_GASV_29[i][0] = las_29_data.data[j, 9]
                Para_GR[i][0] = las_29_data.data[j, 10]
                Para_RHOB[i][0] = las_29_data.data[j, 11]
                Para_CNL[i][0] = las_29_data.data[j, 13]
    data_29 = np.hstack((Para_DTS, Para_DT, Para_RHOB, Para_GR, Aim_29, Para_GASV_29))
    print(data_29.shape)
    
    workbook_35 = openpyxl.load_workbook('./井/WY35.xlsx')
    worksheet_35 = workbook_35.get_sheet_by_name("GASV")
    fileName_35 = './井/WY35.las'
    las_35 = lasio.read(fileName_35)
    Aim_35 = np.array([item.value for item in list(worksheet_35.columns)[2]]).reshape(-1, 1)
    df_Aim_35 = pd.DataFrame(Aim_35, columns=['GASV'])
    las_35_data_depth = las_35.data[:, 0]
    las_35_data = las_35.data[:, ]
    Para_DTS = np.zeros((Aim_35.shape[0], 1))
    Para_RD = np.zeros((Aim_35.shape[0], 1))
    Para_DT = np.zeros((Aim_35.shape[0], 1))
    Para_GASV_35 = np.zeros((Aim_35.shape[0], 1))
    Para_GR = np.zeros((Aim_35.shape[0], 1))
    Para_RHOB = np.zeros((Aim_35.shape[0], 1))
    Para_CNL = np.zeros((Aim_35.shape[0], 1))
    depth_max = np.array([item.value for item in list(worksheet_35.columns)[1]])
    depth = ((depth_max * 10).astype(int)) / 10
    for i in range(Aim_35.shape[0]):
        for j in range(las_35_data_depth.shape[0]):
            if (depth[i] == las_35_data_depth[j]):
                Para_DTS[i][0] = las_35_data.data[j, 3]
                Para_RD[i][0] = las_35_data.data[j, 5]
                Para_DT[i][0] = las_35_data.data[j, 7]
                Para_GASV_35[i][0] = las_35_data.data[j, 9]
                Para_GR[i][0] = las_35_data.data[j, 10]
                Para_RHOB[i][0] = las_35_data.data[j, 11]
                Para_CNL[i][0] = las_35_data.data[j, 13]
    data_35 = np.hstack((Para_DTS, Para_DT, Para_RHOB, Para_GR, Aim_35, Para_GASV_35))
    print(data_35.shape)

    workbook_11 = openpyxl.load_workbook('./井/WY11.xlsx')
    worksheet_11 = workbook_11.get_sheet_by_name("GASV")
    fileName_11 = './井/weiye11jgw.las'
    las_11 = lasio.read(fileName_11)
    Aim_11 = (np.array([item.value for item in list(worksheet_11.columns)[2]]).reshape(-1, 1)) + np.array([item.value for item in list(worksheet_11.columns)[3]]).reshape(-1, 1)
    df_Aim_11 = pd.DataFrame(Aim_11, columns=['GASV'])
    las_11_data_depth = las_11.data[:, 0]
    las_11_data = las_11.data[:, ]
    Para_DTS = np.zeros((Aim_11.shape[0], 1))
    Para_RD = np.zeros((Aim_11.shape[0], 1))
    Para_DT = np.zeros((Aim_11.shape[0], 1))
    Para_GASV_11 = np.zeros((Aim_11.shape[0], 1))
    Para_GR = np.zeros((Aim_11.shape[0], 1))
    Para_RHOB = np.zeros((Aim_11.shape[0], 1))
    Para_CNL = np.zeros((Aim_11.shape[0], 1))
    depth_max = np.array([item.value for item in list(worksheet_11.columns)[1]])
    depth = ((depth_max * 10).astype(int)) / 10
    for i in range(Aim_11.shape[0]):
        for j in range(las_11_data_depth.shape[0]):
            if (depth[i] == las_11_data_depth[j]):
                Para_DTS[i][0] = las_11_data.data[j, 3]
                Para_RD[i][0] = las_11_data.data[j, 5]
                Para_DT[i][0] = las_11_data.data[j, 7]
                Para_GASV_11[i][0] = las_11_data.data[j, 9]
                Para_GR[i][0] = las_11_data.data[j, 10]
                Para_RHOB[i][0] = las_11_data.data[j, 11]
                Para_CNL[i][0] = las_11_data.data[j, 13]
    data_11 = np.hstack((Para_DTS, Para_DT, Para_RHOB, Para_GR, Aim_11, Para_GASV_11))
    print(data_11.shape)


    workbook_23 = openpyxl.load_workbook('./井/WY23.xlsx')
    worksheet_23 = workbook_23.get_sheet_by_name("GASV")
    fileName_23 = './井/WY23.las'
    las_23 = lasio.read(fileName_23)
    Aim_23 = (np.array([item.value for item in list(worksheet_23.columns)[2]]).reshape(-1, 1)) + np.array([item.value for item in list(worksheet_23.columns)[3]]).reshape(-1, 1)
    df_Aim_23 = pd.DataFrame(Aim_23, columns=['GASV'])
    las_23_data_depth = las_23.data[:, 0]
    las_23_data = las_23.data[:, ]
    Para_DTS = np.zeros((Aim_23.shape[0], 1))
    Para_RD = np.zeros((Aim_23.shape[0], 1))
    Para_DT = np.zeros((Aim_23.shape[0], 1))
    Para_GASV_23 = np.zeros((Aim_23.shape[0], 1))
    Para_GR = np.zeros((Aim_23.shape[0], 1))
    Para_RHOB = np.zeros((Aim_23.shape[0], 1))
    Para_CNL = np.zeros((Aim_23.shape[0], 1))
    depth_max = np.array([item.value for item in list(worksheet_23.columns)[1]])
    depth = ((depth_max * 10).astype(int)) / 10
    for i in range(Aim_23.shape[0]):
        for j in range(las_23_data_depth.shape[0]):
            if (depth[i] == las_23_data_depth[j]):
                Para_DTS[i][0] = las_23_data.data[j, 3]
                Para_RD[i][0] = las_23_data.data[j, 5]
                Para_DT[i][0] = las_23_data.data[j, 7]
                Para_GASV_23[i][0] = las_23_data.data[j, 9]
                Para_GR[i][0] = las_23_data.data[j, 10]
                Para_RHOB[i][0] = las_23_data.data[j, 11]
                Para_CNL[i][0] = las_23_data.data[j, 13]
    data_23 = np.hstack((Para_DTS, Para_DT, Para_RHOB, Para_GR, Aim_23, Para_GASV_23))
    print(data_23.shape)
    data = np.vstack((data_1, data_29, data_35, data_11, data_23))
    return data